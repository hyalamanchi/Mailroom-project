#!/usr/bin/env python3
"""
Case Review File Generator with Google Drive Upload Capability
Uses service account authentication for production stability.
"""

import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import logging

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Google Drive API scopes
SCOPES = ['https://www.googleapis.com/auth/drive.file']

class CaseReviewer:
    def __init__(self, input_file):
        self.input_file = input_file
        self.output_dir = os.path.join(os.path.dirname(input_file), "REVIEWED_CASES")
        os.makedirs(self.output_dir, exist_ok=True)
        self.service = None
        
    def get_drive_service(self):
        """Set up Google Drive API service using service account"""
        if self.service:
            return self.service
            
        try:
            creds = service_account.Credentials.from_service_account_file(
                'service-account-key.json',
                scopes=SCOPES
            )
            self.service = build('drive', 'v3', credentials=creds)
            logger.info("✅ Authenticated with Google Drive (Service Account)")
            return self.service
            
        except FileNotFoundError:
            logger.error("❌ Error: service-account-key.json not found")
            logger.error("💡 Place your service account JSON file in the project root")
            raise
        except Exception as e:
            logger.error(f"❌ Authentication error: {str(e)}")
            raise

    def upload_to_drive(self, file_path, folder_id=None):
        """Upload file to Google Drive"""
        try:
            service = self.get_drive_service()
            
            file_metadata = {
                'name': os.path.basename(file_path),
                'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            if folder_id:
                file_metadata['parents'] = [folder_id]
            
            media = MediaFileUpload(
                file_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                resumable=True
            )
            
            file = service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id, webViewLink'
            ).execute()
            
            logger.info(f"✅ File uploaded to Drive: {file.get('webViewLink')}")
            return file.get('webViewLink')
            
        except Exception as e:
            logger.error(f"❌ Error uploading to Drive: {str(e)}")
            raise

    def fix_notice_date(self, df):
        """Fix notice dates to be in current year"""
        try:
            # Convert notice_date to datetime if it's not already
            df['notice_date'] = pd.to_datetime(df['notice_date'], errors='coerce')
            
            # Get current year
            current_year = datetime.now().year
            
            # Identify dates after current year
            future_dates = df['notice_date'] > f'{current_year}-12-31'
            
            # For future dates, adjust to current year
            df.loc[future_dates, 'notice_date'] = df.loc[future_dates, 'notice_date'].apply(
                lambda x: x.replace(year=current_year)
            )
            
            # Format dates as string in MM/DD/YYYY format
            df['notice_date'] = df['notice_date'].dt.strftime('%m/%d/%Y')
            
            return df
            
        except Exception as e:
            logger.error(f"Error fixing notice dates: {str(e)}")
            return df

    def process_cases(self):
        """Process the cases Excel file and add action buttons"""
        try:
            # Read the Excel file
            logger.info(f"Reading input file: {self.input_file}")
            df = pd.read_excel(self.input_file)
            
            # Fix notice dates
            logger.info("Fixing notice dates...")
            df = self.fix_notice_date(df)
            
            # Initialize tracking columns
            tracking_columns = ['Status', 'Review Notes', 'Last Updated', 'Reviewed By']
            for col in tracking_columns:
                if col not in df.columns:
                    df[col] = 'Pending' if col == 'Status' else ''
            
            # Split cases
            matched_cases = df[df['is_matched'] if 'is_matched' in df.columns else df['case_id'].notna()].copy()
            unmatched_cases = df[~df['is_matched'] if 'is_matched' in df.columns else df['case_id'].isna()].copy()
            logger.info(f"Found {len(matched_cases)} matched cases and {len(unmatched_cases)} unmatched cases")
            
            # Generate output filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = os.path.join(
                self.output_dir, 
                f"CASES_WITH_ACTIONS_{timestamp}.xlsx"
            )
            
            # Create Excel writer to save to multiple sheets
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                matched_cases.to_excel(writer, sheet_name='Matched Cases', index=False)
                unmatched_cases.to_excel(writer, sheet_name='Unmatched Cases', index=False)
            
            # Load workbook for formatting
            wb = load_workbook(output_file)
            
            # Create button style once for the workbook
            try:
                button_style = wb.named_styles['button_style']
            except KeyError:
                button_style = NamedStyle(name='button_style')
                button_style.font = Font(bold=True)
                button_style.fill = PatternFill(
                    start_color='4472C4',
                    end_color='4472C4',
                    fill_type='solid'
                )
                wb.add_named_style(button_style)
            
            # Format each sheet
            for sheet_name in ['Matched Cases', 'Unmatched Cases']:
                ws = wb[sheet_name]
                
                # Insert summary row at the top
                ws.insert_rows(1)
                summary_cell = ws.cell(row=1, column=1)
                if sheet_name == 'Matched Cases':
                    summary_cell.value = f"✅ Matched Cases ({ws.max_row - 2} cases)"
                    summary_cell.fill = PatternFill(
                        start_color='C6EFCE',  # Light green
                        end_color='C6EFCE',
                        fill_type='solid'
                    )
                else:
                    summary_cell.value = f"⚠️ Unmatched Cases ({ws.max_row - 2} cases)"
                    summary_cell.fill = PatternFill(
                        start_color='FFC7CE',  # Light red
                        end_color='FFC7CE',
                        fill_type='solid'
                    )
                summary_cell.font = Font(bold=True)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(list(ws.columns)))
                
                # Format headers in row 2
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=2, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color='CCCCCC',
                        end_color='CCCCCC',
                        fill_type='solid'
                    )
                    
                    # Find Status column while iterating
                    if cell.value == 'Status':
                        status_col = col
                
                # Set up data validation for Status column
                dv = DataValidation(
                    type="list",
                    formula1='"Approve,Reject,Review,Pending"',
                    allow_blank=False
                )
                ws.add_data_validation(dv)
                
                # Apply validation and formatting to Status column
                for row in range(3, ws.max_row + 1):  # Start from row 3 (data rows)
                    cell = ws.cell(row=row, column=status_col)
                    
                    # Add validation to cell
                    dv.add(cell)
                    
                    # Set default value for empty cells
                    if not cell.value:
                        cell.value = 'Pending'
                    
                    # Apply color coding
                    status_colors = {
                        'Pending': 'FFEB9C',  # Light yellow
                        'Approve': 'C6EFCE',  # Light green
                        'Reject': 'FFC7CE',   # Light red
                        'Review': 'BDD7EE'    # Light blue
                    }
                    
                    if cell.value in status_colors:
                        cell.fill = PatternFill(
                            start_color=status_colors[cell.value],
                            end_color=status_colors[cell.value],
                            fill_type='solid'
                        )
                
                # Adjust column widths
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width
            
            # Save the formatted file
            wb.save(output_file)
            logger.info(f"✅ Successfully created review file: {output_file}")
            
            # Upload to Google Drive
            drive_link = self.upload_to_drive(output_file)
            
            return output_file, drive_link
            
        except Exception as e:
            logger.error(f"❌ Error processing cases: {str(e)}")
            raise

def process_and_upload(matched_file=None, unmatched_file=None):
    """Process and upload review files"""
    try:
        # Default files if not provided
        if not matched_file:
            matched_file = os.path.join("QUALITY_REVIEW", "QUALITY_REVIEW_MATCHED_CASES.xlsx")
        if not unmatched_file:
            unmatched_file = os.path.join("QUALITY_REVIEW", "UNMATCHED_CASES.xlsx")
        
        # Create a combined output folder
        base_dir = os.path.dirname(matched_file)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        combined_output = os.path.join(base_dir, f"COMBINED_REVIEW_{timestamp}.xlsx")
        
        # Read both files
        logger.info("Reading matched cases file...")
        matched_df = pd.read_excel(matched_file)
        logger.info("Reading unmatched cases file...")
        unmatched_df = pd.read_excel(unmatched_file)
        
        # Add tracking columns to both
        tracking_columns = ['Status', 'Review Notes', 'Last Updated', 'Reviewed By']
        for df in [matched_df, unmatched_df]:
            for col in tracking_columns:
                if col not in df.columns:
                    df[col] = 'Pending' if col == 'Status' else ''
        
        logger.info("Creating combined Excel file...")
        # Save to combined Excel file
        with pd.ExcelWriter(combined_output, engine='openpyxl') as writer:
            matched_df.to_excel(writer, sheet_name='Matched Cases', index=False)
            unmatched_df.to_excel(writer, sheet_name='Unmatched Cases', index=False)
        
        # Format the combined file and upload to Drive
        logger.info("Formatting and uploading to Drive...")
        reviewer = CaseReviewer(combined_output)
        output_file, drive_link = reviewer.process_cases()
        
        print("\n=== Case Review File Created and Uploaded ===")
        print(f"Matched file: {matched_file}")
        print(f"Unmatched file: {unmatched_file}")
        print(f"Combined output: {output_file}")
        print(f"Google Drive link: {drive_link}")
        print("\nInstructions:")
        print("1. Open the file using the Google Drive link")
        print("2. For each case:")
        print("   - Click the Status cell to see the action buttons")
        print("   - Select an action for each case (Approve/Reject/Review)")
        print("   - Add any notes in the Review Notes column")
        print("   - Your name will be automatically added to Reviewed By")
        print("3. The file will auto-save in Google Drive")
        
        return output_file, drive_link
        
    except Exception as e:
        logger.error(f"❌ Error: {str(e)}")
        raise

def main():
    try:
        process_and_upload()
        return 0
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        return 1

if __name__ == "__main__":
    exit(main())